# -*- coding: utf-8 -*-
"""
Created on Thu Dec 10 10:02:29 2020

@author: Erinb
"""

# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
from openpyxl import load_workbook  # pip install openpyxl
#import datetime
#from contextlib import contextmanager
from pprint import pprint # pretty print
from expt import bounding_box_8

class Annotations:
   """read in an annotation sheet
   .heading[name] -> col num
   .Labels:set avaiable after getLines has been run
   """
   # how to make a instance of the class
   def __init__(self, file_name:str):  # __init__ is a reserved name to instantiate a class
     #print(file_name)
      wb = load_workbook(file_name)
      self._book = wb
      self.FileName = file_name
      
      ws = wb.active
      self._sheet = ws
      self._header = h =  dict((name.value, # don't want the formula or  just the cell object
                                col # enumerate is 0-based - col A is 1
                                ) for (col, name) in enumerate(ws[1]))
      pprint(h, sort_dicts=False) # not by label name, but col number

   def getLines(self, header_rows=1, number_frames=True):
      h = self._header

      filenames = set(x.value for x in self._sheet['I'])# FIX ME - should be h['filename']. probably chr(ord('A')+h['filename'])
      filenames.remove('filename') #remove extra column filename if present 
      self.FileNames = filenames #produce txt. file
      frame_nums = dict((fn, i) for i, fn in enumerate(sorted(filenames)))
      print("\n Frame num map") #\n is new line command for the operating system that is being used
      pprint(frame_nums, sort_dicts=False) # not by label name, but col number

      labels_seen = set()                                        # to become labels.txt
      
      for row in self._sheet.iter_rows(min_row = header_rows+1): # skip the header row
         label_name = row[h["label_name"]].value
         if label_name.lower() == "laser point":
            print("skip laser point")
            continue
         
         p = tuple(map(float, row[h["points"]].value.strip()[1:-1].split(',')))           # remove whitespace, then [ and ]
     #    print("Points", p)
         shape_name = row[h["shape_name"]].value.lower()
         if shape_name == "circle":
            x, y, r = p
            tl_x = x-r
            tl_y = y-r
            br_x = x+r
            br_y = y+r
         elif shape_name == "ellipse": # think it's maj A, min A, maj B, min B
            # just go for simple bounding box (may be too small!)
            x = p[::2]          # odd ones 
            y = p[1::2]         # even ones
            if False: #change to 'True' to get to old case 
               tl_x = min(x)
               tl_y = min(y)
               br_x = max(x)
               br_y = max(y)
            else:
             (tl_x, tl_y),(br_x, br_y) = bounding_box_8(p) #function for ellipse
         elif shape_name == "point": # this is doesn't have a  right size so it a bit crappy code
            x, y =  p
            r=1                 # arbitrary - but zero feels bad
            tl_x = x-r
            tl_y = y-r
            br_x = x+r
            br_y = y+r
         elif shape_name == "linestring":
            x = p[::2]
            y = p[1::2]
            r=1
            tl_x = min(x)
            tl_y = min(y)
            br_x = max(x)
            br_y = max(y)
         else:
            print(f"Don't understand shape type {shape_name}")#  f = formatted or b = bytes (alphabet)
            raise ValueError(shape_name, p)

         # for readability, create the values as int, float, ... then turn to string in a single map
         
         yield ",".join(map(str, (row[h["annotation_label_id"]].value, # join takes one argument
                        this_fn := row[h["filename"]].value,
                        frame_nums[this_fn] if number_frames else 0,    # frame id - 0 in a single image...
                        int(tl_x),#tl is top left
                        int(tl_y),
                        int(br_x),#br is bottom right
                        int(br_y),
                        0.99, # prob correct
                        -1,   # length
                        label_name,
                        0.99)))

         labels_seen.add(label_name)
      self.Labels = labels_seen
         
if __name__ == "__main__":  # only executes if the file is being run directly from command line
   from pathlib import Path
   
   stuff = Annotations(p:= Path(r"D:\Research_Masters\ROV_data\ADS_800m\ADS_800m_T3_Framegrabs_1min-int.xlsx"))
   
   #.replace('\\', '/') = changes slashes
   #r = raw string > solves back-slash issue
   # using with guarantees the file will be closed - even with errors
   with open(p.with_suffix(".csv"), "wt") as f: # wt - write text
      for i, r in enumerate(stuff.getLines()): # enumerate return (seq_num, thing)
      #   print(i,r)
       #  print()
         #if i > 10:
         #   break
         print(r, file=f)
#create OTU labels - unsure what this is used for possibly yolo?
   with open(p.parent / "labels.txt", "wt") as f: # .parent is the folder. / is 'overloaded' to 
                                                  # join up elements of a path
      for l in sorted(stuff.Labels, key = lambda x: int(x[3:]) if x.startswith('OTU') else 9999): 
         print(l, file=f)
#create the txt. file with the list of image names for VIAME
   with open(p.parent / "filenames_T3.txt", "wt") as f: # .parent is the folder. / is 'overloaded' to 
                                                  # join up elements of a path
      for l in sorted(stuff.FileNames): 
         print(l, file=f)
