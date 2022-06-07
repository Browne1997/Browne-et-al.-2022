# -*- coding: utf-8 -*-
"""
Code to run over the annotations in a ?BIIGLE? xls, create the VIAME csv, and also
create directories containing the image clips - one directory per OTU
"""

# -*- coding: utf-8 -*-

from openpyxl import load_workbook  # pip install openpyxl
#import datetime
#from contextlib import contextmanager
from pprint import pprint # pretty print
from expt import bounding_box_8
from pathlib import Path
import csv, re

# to allow control over from xls2csv import *
__all__ = ("Annotations", "AnnotationsCSV", "VAnnotationsCSV", "Place", "ViameCSV")


class Place:
   """Information for a Transect or other period of diving
   """
   _places = {}                 # string to Place

   # re, optional, type to convert to
   
   REs = { 'Site':     (re.compile(r"""((?:ADS)|  # (?: groups characters, but doesn't create a group
                                        (?:NRB)|  # comments are allowed!
                                        (?:RB))""", re.VERBOSE | re.I), False, None),    # one of the sites - case independent
           'Depth':    (re.compile(r"""(\d+)m?""", re.VERBOSE | re.I), False, int), #  decimal, optional m
           'Transect': (re.compile(r"""T(\d)"""), False, int),
           'Time':     (re.compile(r"""([012]\d-[0-5]\d-[0-5]\d)"""), True, None),
           }

   rest_num_RE = re.compile(r'(\d+)')
   
   DiveNames = set()
   _rests = set()
   
   @classmethod
   def findPlace(cls, s):
      s = str(Path(s).stem)
      try:
         return cls._places[s]
      except KeyError:
         p = cls(s)
         cls._places[s] = p
         return p
      
   def __init__(self, s:str):
      """
      4 properties for a place:
      - Depth: int
      - Transect: int
      - Time: str
      - Other: str
      """

      s = str(Path(s).stem)

      self.FileName = s

      # print("Place for ", s, end=' ')
      
      pos = 0
      
      for name, (RE, optional, type_) in self.REs.items():
         #print("Name", name)
         match = RE.search(s, pos)

         if optional and not match:
            setattr(self, name, None)
            continue
         
         assert match
         
         setattr(self, name, (type_(match[1]) if type_ else match[1]))
         pos = match.end(1)+1

      self._rest = s[pos:]

      match_in_rest = self.rest_num_RE.search(self._rest)

      num_in_rest = f"_{match_in_rest[1]}" if match_in_rest else ""

      dn = f"{self.Depth}_T{self.Transect}_{self.Site}{num_in_rest}"

      #if num_in_rest:
      #   print("NIR", num_in_rest, dn, s)
         
      self.DiveName = dn
      #print(self)
      
   def __str__(self):
      return(f"Site {self.Site} Depth {self.Depth}m T{self.Transect} @ {self.Time} + {self._rest}")


   def __eq__(self, other):
      return self.Site == other.Site and self.Depth == other.Depth and self.Transect == other.Transect

class ViameCSV:
   @classmethod
   def Header(cls, leading_comment = True):
      return (f"{'# ' if leading_comment else ''}annotation_label_id", 
              "filename",
              "frame_number",
              "top_left_x",
              "top_left_y",
              "bottom_right_x",
              "bottom_right_y",
              "prob_bounds",
              "length",
              "label_name",
              "prob_label")
   
class Annotations(ViameCSV):
   """read in an annotation sheet
   .heading[name] -> col num
   .Labels:set available after getLines has been run
   """
   # how to make a instance of the class
   def __init__(self, file_name):  # __init__ is a reserved name to instantiate a class
     #print(file_name)
      file_name = Path(file_name)
      wb = load_workbook(str(file_name)) # can be a Path
      self._book = wb
      self.FileName = file_name
      
      ws = wb.active
      self._sheet = ws
      self._header = h =  dict((name.value, # don't want the formula or  just the cell object
                                col # enumerate is 0-based - col A is 1
                                ) for (col, name) in enumerate(ws[1]))
      #pprint(h, sort_dicts=False) # not by label name, but col number

      assert (fn:= self._sheet['I'][0].value) == "filename", f"Col I isn't filename but {fn}"

      filenames = set(x.value for x in self._sheet['I'][1:]) # skip header row

      self.FileNames = filenames # need for Viame 

      self.Place = Place.findPlace(file_name.stem)

   def getLines(self, *, header_rows=1, number_frames=True, with_header=True):    # after * must be given by keyword
      """ header_rows are to be skipped to reach the first annotation record
          number_frames True sets the frame number to the relevant still image (Viame treats dir of image like a video)

          yields a series of 9-tuples, as required by Viame. NB not converted to str

          after completion, self.Labels is a set of all the OTUs seen
      """
      
      h = self._header
      filenames = self.FileNames
      
      frame_nums = dict((fn, i) for i, fn in enumerate(sorted(filenames)))
      #print("\n Frame num map") #\n is new line command for the operating system that is being used
      #pprint(frame_nums, sort_dicts=False) # not by label name, but col number

      labels_seen = set()                                        # to become labels.txt
      if with_header:
         yield self.Header(), self.Place
         
      for row in self._sheet.iter_rows(min_row = header_rows+1): # skip the header row
         label_name = row[h["label_name"]].value
         if label_name.lower() == "laser point":
            # print("skip laser point")
            continue
         
         p = tuple(map(float, row[h["points"]].value.strip()[1:-1].split(',')))           # remove whitespace, then [ and ]
     #    print("Points", p)
         shape_name = row[h["shape_name"]].value.lower()
         if shape_name == "circle":
            x, y, r = p
            tl_x = x-r
            tl_y = y-r
            br_x = x+r
            br_y = y+r
         elif shape_name == "ellipse": # think it's maj A, min A, maj B, min B
            # just go for simple bounding box (may be too small!)
            x = p[::2]          # odd ones 
            y = p[1::2]         # even ones
            if False: #change to 'True' to get to old case 
               tl_x = min(x)
               tl_y = min(y)
               br_x = max(x)
               br_y = max(y)
            else:
             (tl_x, tl_y),(br_x, br_y) = bounding_box_8(p) #function for ellipse
         elif shape_name == "point": # this is doesn't have a  right size so it a bit crappy code
            x, y =  p
            r=1                 # arbitrary - but zero feels bad
            tl_x = x-r
            tl_y = y-r
            br_x = x+r
            br_y = y+r
         elif (shape_name == "linestring") or (shape_name == "polygon"):
            x = p[::2]
            y = p[1::2]
            tl_x = min(x)
            tl_y = min(y)
            br_x = max(x)
            br_y = max(y)
         else:
            print(f"Don't understand shape type {shape_name}")#  f = formatted or b = bytes (alphabet)
            raise ValueError(shape_name, p)

         # for readability, create the values as int, float, ... then turn to string in a single map
         
         yield (row[h["annotation_label_id"]].value, # join takes one argument, so yield a tuple
                this_fn := row[h["filename"]].value,
                frame_nums[this_fn] if number_frames else 0,    # frame id - 0 in a single image...
                int(tl_x),#tl is top left
                int(tl_y),
                int(br_x),#br is bottom right
                int(br_y),
                0.99, # prob correct
                -1,   # length
                label_name,
                0.99), Place.findPlace(Path(this_fn))
         
         labels_seen.add(label_name)
      self.Labels = labels_seen
         

class AnnotationsCSV(ViameCSV):
   """read in an annotation sheet
   .heading[name] -> col num
   .Labels:set available after getLines has been run
   """
   # how to make a instance of the class
   def __init__(self, file_name:str):  # __init__ is a reserved name to instantiate a class
      #print(file_name)
      assert Path(file_name).suffix == '.csv', file_name
     
      self.FileName = file_name

      try:
         self.Place = Place.findPlace(file_name)
      except:
         self.Place = None
         
      got_hdr = False
      self._rows = []
      
      with open(file_name, "rt", newline='') as f:
         for fields in csv.reader(f): 
            if not fields: # skip blanks lines
               continue
            
            # for viame, # marks a comment line in their csv, and they don't understand a header line
            # excel probably won't like #, but does have header lines...
            # so use #field1,field2,
            
            if not got_hdr:
               #print("header stuff", fields)
               try:
                  int(fields[0])   # if the first field is an int, then the header line is missing...
               except:
                  pass
               else:
                  raise ValueError("Header line missing", l)
               self._header = h = dict((name, col) for (col, name) in enumerate(fields))
               key1, val1 = list(h.items())[0]
               if key1.startswith('#'):
                  del h[key1]
                  h[key1[1:].strip()] = val1   #  chop off the #, then any leading whitespace
               got_hdr = True
            else:
               try:
                  if fields[0].startswith('#'):
                     continue
               except:
                  pass
               self._rows.append(fields)
               
      # pprint(h, sort_dicts=False) # not by label name, but col number

      filename_col = h['filename']
      
      filenames = set(x[filename_col] for x in self._rows)

      self.FileNames = filenames # need for Viame 

      
   def getLines(self, *, header_rows=1, number_frames=True, with_header=True):    # after * must be given by keyword
      """ header_rows are to be skipped to reach the first annotation record
          number_frames True sets the frame number to the relevant still image (Viame treats dir of image like a video)

          yields a series of 9-tuples, as required by Viame. NB not converted to str

          after completion, self.Labels is a set of all the OTUs seen
      """
      
      h = self._header
      filenames = self.FileNames
      
      frame_nums = dict((fn, i) for i, fn in enumerate(sorted(filenames)))
      #print("\n Frame num map") #\n is new line command for the operating system that is being used
      #pprint(frame_nums, sort_dicts=False) # not by label name, but col number

      labels_seen = set()                                        # to become labels.txt
      if with_header:
         yield self.Header(), self.Place

      max_min_shapes = {"linestring", "polygon", "rectangle"}
      
      for row_num, row in enumerate(self._rows):
         try:
            label_name = row[h["label_name"]]
            if label_name.lower() == "laser point":
               # print("skip laser point")
               continue
            
            p = tuple(map(float, row[h["points"]][1:-1].split(',')))           # remove whitespace, then [ and ]
            # print("Points", p, row[h["points"]])
            shape_name = row[h["shape_name"]].lower()
            if shape_name == "circle":
               x, y, r = p
               tl_x = x-r
               tl_y = y-r
               br_x = x+r
               br_y = y+r
            elif shape_name == "ellipse": # think it's maj A, min A, maj B, min B
               # just go for simple bounding box (may be too small!)
               x = p[::2]          # odd ones 
               y = p[1::2]         # even ones
               (tl_x, tl_y),(br_x, br_y) = bounding_box_8(p) #function for ellipse
            elif shape_name == "point": # this is doesn't have a  right size so it a bit crappy code
               x, y =  p
               r=1                 # arbitrary - but zero feels bad
               tl_x = x-r
               tl_y = y-r
               br_x = x+r
               br_y = y+r
            elif shape_name in max_min_shapes:
               x = p[::2]
               y = p[1::2]
               tl_x = min(x)
               tl_y = min(y)
               br_x = max(x)
               br_y = max(y)
            else:
               print(f"Don't understand shape type {shape_name}")#  f = formatted or b = bytes (alphabet)
               raise ValueError(shape_name, p)
   
            # for readability, create the values as int, float, ... then turn to string in a single map
            
            yield (int(row[h["annotation_label_id"]]), # join takes one argument, so yield a tuple
                   (this_fn := row[h["filename"]]),
                   frame_nums[this_fn] if number_frames else 0,    # frame id - 0 in a single image...
                   int(tl_x),#tl is top left
                   int(tl_y),
                   int(br_x),#br is bottom right
                   int(br_y),
                   0.99, # prob correct
                   -1,   # length
                   label_name,
                   0.99), Place.findPlace(Path(this_fn))
            
            labels_seen.add(label_name)
         except Exception as e:
            print(f"{row_num+2}: {row =} failed {e}")
            raise
      self.Labels = labels_seen
         
class VAnnotationsCSV(AnnotationsCSV):
   def getLines(self, *, header_rows=1, number_frames=True, with_header=True):    # after * must be given by keyword
      """ header_rows are to be skipped to reach the first annotation record
          number_frames True sets the frame number to the relevant still image (Viame treats dir of image like a video)

          yields a series of 9-tuples, as required by Viame. NB not converted to str

          after completion, self.Labels is a set of all the OTUs seen
      """
      
      h = self._header
      filenames = self.FileNames
      
      frame_nums = dict((fn, i) for i, fn in enumerate(sorted(filenames)))
      #print("\n Frame num map") #\n is new line command for the operating system that is being used
      #pprint(frame_nums, sort_dicts=False) # not by label name, but col number

      labels_seen = set()                                        # to become labels.txt
      if with_header:
         yield self.Header(), None #self.Place

      max_min_shapes = {"linestring", "polygon", "rectangle"}
      
      for row_num, row in enumerate(self._rows):
         try:
               
            yield (int(row[0]), # join takes one argument, so yield a tuple
                   (this_fn := row[1]),
                   int(row[2]) if number_frames else 0,    # frame id - 0 in a single image...
                   int(row[3]), #tl is top left
                   int(row[4]),
                   int(row[5]),#br is bottom right
                   int(row[6]),
                   float(row[7]), # prob correct
                   -1,   # length
                   (label_name := row[9]),
                   float(row[10])), None  # Place.findPlace(Path(this_fn))
            
            labels_seen.add(label_name)
         except Exception as e:
            print(f"{row_num+2}: {row =} failed {e}")
            raise
      self.Labels = labels_seen
         

if __name__ == "__main__":  # only executes if the file is being run directly from command line
   
   import argparse
   # construct the argument parse and parse the arguments
   ap = argparse.ArgumentParser(description="Convert Viame human annotations for use in Viame")
   
   #ap.add_argument("-i", "--interval", type=float, default=2.0,
   #                help="time between frames (s)")
   ap.add_argument("-c", "--csv", action="store_true", 
                   help="convert from the CSV form of Biigle, not xls")
   
   ap.add_argument("input_name",
                   help="input spreadsheet")

   args = ap.parse_args()
   
   p = Path(args.input_name)
   assert p.suffix == (".xlsx" if not args.csv  else ".csv") 
   
   # *rest - for now don't mind how many other fields (separated by _)

   # where, depth_str, transect, *rest = p.stem.replace('-', '_').split('_'a)
   
   stuff = (AnnotationsCSV if args.csv else Annotations)(p)
   
   #.replace('\\', '/') = changes slashes
   #r = raw string > solves back-slash issue
   # using with guarantees the file will be closed - even with errors

   top_place = stuff.Place
   
   with open(f"{p.stem}_v.csv", "wt") as f: # wt - write text

      print(f"Creating VIAME csv in {f.name}")

      for i, (r, place) in enumerate(stuff.getLines()): # enumerate return (seq_num, thing)

         print(",".join(map(str, r)), file=f)

         assert place == top_place, f"{place} not in {top_place}"
         
   # create OTU labels - unsure what this is used for possibly yolo?
   # djch - used for yolo training
   
   with open(p.parent / f"{p.stem}_labels.txt", "wt") as f: # .parent is the folder. / is 'overloaded' to 
                                                  # join up elements of a path
      print(f"Creating one file labels in {f.name}")
      for l in sorted(stuff.Labels, key = lambda x: int(x[3:]) if x.startswith('OTU') else 9999): 
         print(l, file=f)

   # create the txt. file with the list of image names for VIAME
   with open(p.parent / f"filenames_{top_place.Site}_{top_place.Depth}_T{top_place.Transect}_{top_place._rest}.txt", "wt") as f: # .parent is the folder. / is 'overloaded' to 
                                                        # join up elements of a path
      print(f"Creating filename list in {f.name}")
      for l in sorted(stuff.FileNames): 
         print(l, file=f)
